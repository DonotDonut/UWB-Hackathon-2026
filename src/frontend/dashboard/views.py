from django.shortcuts import render, redirect
from openpyxl import Workbook, load_workbook
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parents[3]
EXCEL_FILE = BASE_DIR / "data" / "staff_data.xlsx"

def dashboard(request):
    return render(request, "dashboard.html")

def add_person(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        role = request.POST.get("role")

        if EXCEL_FILE.exists():
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Staff"
            sheet.append(["Full Name", "Email", "Role", "", "", ""])

        sheet.append([full_name, email, role])
        workbook.save(EXCEL_FILE)

        return redirect("add_person")

    return render(request, "add_person.html")

def schedules(request):
    staff_schedule = [
        {"employee": "John Smith", "days": "Monday, Tuesday, Wednesday, Friday", "time": "9:00 AM - 5:00 PM"},
        {"employee": "Jane Doe", "days": "Tuesday, Wednesday, Thursday", "time": "12:00 PM - 8:00 PM"},
        {"employee": "Mike Johnson", "days": "Monday, Wednesday, Friday", "time": "9:00 AM - 5:00 PM"},
        {"employee": "Emily Davis", "days": "Thursday, Friday, Saturday", "time": "2:00 PM - 10:00 PM"},
        {"employee": "Chris Lee", "days": "Monday, Tuesday, Friday", "time": "9:00 AM - 5:00 PM"},
    ]

    week_schedule = [
        {
            "day": "Monday",
            "shifts": [
                {"employee": "John Smith", "time": "9:00 AM - 5:00 PM", "color": "john"},
                {"employee": "Mike Johnson", "time": "9:00 AM - 5:00 PM", "color": "mike"},
            ],
        },
        {
            "day": "Tuesday",
            "shifts": [
                {"employee": "John Smith", "time": "9:00 AM - 5:00 PM", "color": "john"},
                {"employee": "Jane Doe", "time": "12:00 PM - 8:00 PM", "color": "jane"},
                {"employee": "Chris Lee", "time": "9:00 AM - 5:00 PM", "color": "chris"},
            ],
        },
        {
            "day": "Wednesday",
            "shifts": [
                {"employee": "John Smith", "time": "9:00 AM - 5:00 PM", "color": "john"},
                {"employee": "Mike Johnson", "time": "9:00 AM - 5:00 PM", "color": "mike"},
                {"employee": "Jane Doe", "time": "12:00 PM - 8:00 PM", "color": "jane"},
            ],
        },
        {
            "day": "Thursday",
            "shifts": [
                {"employee": "Jane Doe", "time": "12:00 PM - 8:00 PM", "color": "jane"},
                {"employee": "Emily Davis", "time": "2:00 PM - 10:00 PM", "color": "emily"},
            ],
        },
        {
            "day": "Friday",
            "shifts": [
                {"employee": "John Smith", "time": "9:00 AM - 5:00 PM", "color": "john"},
                {"employee": "Chris Lee", "time": "9:00 AM - 5:00 PM", "color": "chris"},
                {"employee": "Emily Davis", "time": "2:00 PM - 10:00 PM", "color": "emily"},
            ],
        },
        {
            "day": "Saturday",
            "shifts": [
                {"employee": "Emily Davis", "time": "2:00 PM - 10:00 PM", "color": "emily"},
            ],
        },
        {
            "day": "Sunday",
            "shifts": [],
        },
    ]

    return render(
        request,
        "schedules.html",
        {
            "staff_schedule": staff_schedule,
            "week_schedule": week_schedule,
        },
    )

def edit_shift(request):
    employees = []
    shifts = []

    workbook = load_workbook(EXCEL_FILE)

    staff_sheet = workbook["Staff"]

    if "Shifts" not in workbook.sheetnames:
        shifts_sheet = workbook.create_sheet("Shifts")
        shifts_sheet.append(["Employee", "Day", "Start Time", "End Time"])
        workbook.save(EXCEL_FILE)
    else:
        shifts_sheet = workbook["Shifts"]

    # Load employees
    for row in staff_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            employees.append(row[0])

    # Add shift
    if request.method == "POST" and request.POST.get("action") == "add":
        employee_name = request.POST.get("employee")
        days = request.POST.getlist("days")
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")

        for day in days:
            shifts_sheet.append([employee_name, day, start_time, end_time])

        workbook.save(EXCEL_FILE)
        return redirect("edit_shift")

    # Delete shift
    if request.method == "POST" and request.POST.get("action") == "delete":
        row_number = int(request.POST.get("row_number"))
        shifts_sheet.delete_rows(row_number)
        workbook.save(EXCEL_FILE)
        return redirect("edit_shift")

    # Load existing shifts
    for index, row in enumerate(shifts_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:
            shifts.append({
                "row_number": index,
                "employee": row[0],
                "day": row[1],
                "start_time": row[2],
                "end_time": row[3],
            })

    return render(
        request,
        "edit_shift.html",
        {
            "employees": employees,
            "shifts": shifts,
        }
    )