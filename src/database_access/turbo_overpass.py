
# Importing Python Libraries
import os
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class TurboOverpass:
    """
    TurboOverpass API Access and Excel Export Class

    This class is responsible for:
    1. Sending Overpass API queries to OpenStreetMap
    2. Extracting relevant location/store fields from returned OSM elements
    3. Saving the extracted data into a formatted Excel workbook

    The exported data is used later for:
    - Store/restaurant location analysis
    - GIS/KML generation
    - Employee profile generation
    - Store capacity estimation
    """

    # Excel column headers used when exporting Overpass results.
    HEADERS = [
        "osm_type", "osm_id", "name", "lat", "lon",
        "shop", "amenity", "brand", "operator", "phone", "website",
        "addr:housenumber", "addr:street", "addr:city", "addr:state", "addr:postcode",
        "tags_all"
    ]

    @staticmethod
    def fetch_overpass(url, query):
        """
        Send a query to the Overpass API and return the JSON response.

        Parameters:
            url (str): Overpass API endpoint URL.
            query (str): Overpass QL query string.

        Returns:
            dict: JSON response from the Overpass API.

        Raises:
            requests.exceptions.HTTPError:
                Raised when the Overpass API returns a failed status code.
        """

        # Submit Overpass query using POST request.
        resp = requests.post(
            url,
            data={"data": query},
            headers={
                "Accept": "application/json",
                "User-Agent": "IFooDBee-Merchant-Identifier/1.0"
            },
            timeout=180
        )

        # Print detailed error information if the request fails.
        if not resp.ok:
            print("STATUS:", resp.status_code)
            print("RESPONSE TEXT:")
            print(resp.text)
            resp.raise_for_status()

        return resp.json()

    @staticmethod
    def safe_get(tags, key, default=""):
        """
        Safely retrieve a value from an OSM tags dictionary.

        Parameters:
            tags (dict): OSM tag dictionary.
            key (str): Key to retrieve from the dictionary.
            default (str): Default value if key is missing.

        Returns:
            str: Retrieved tag value or default.
        """
        return tags.get(key, default)

    @classmethod
    def extract_row(cls, element):
        """
        Convert a single OpenStreetMap element into an Excel row.

        Parameters:
            element (dict): OSM element returned from Overpass API.

        Returns:
            list: Row values matching the HEADERS order.
        """

        # Extract the tags dictionary safely.
        tags = element.get("tags", {}) if isinstance(element.get("tags", {}), dict) else {}

        # Store all raw tags in one string for traceability/debugging.
        tags_all = "; ".join([f"{k}={v}" for k, v in sorted(tags.items())])

        return [
            element.get("type", ""),
            element.get("id", ""),
            cls.safe_get(tags, "name", ""),
            element.get("lat"),
            element.get("lon"),
            cls.safe_get(tags, "shop", ""),
            cls.safe_get(tags, "amenity", ""),
            cls.safe_get(tags, "brand", ""),
            cls.safe_get(tags, "operator", ""),
            cls.safe_get(tags, "phone", ""),
            cls.safe_get(tags, "website", ""),
            cls.safe_get(tags, "addr:housenumber", ""),
            cls.safe_get(tags, "addr:street", ""),
            cls.safe_get(tags, "addr:city", ""),
            cls.safe_get(tags, "addr:state", ""),
            cls.safe_get(tags, "addr:postcode", ""),
            tags_all
        ]

    @classmethod
    def write_excel(cls, elements, out_path):
        """
        Write Overpass API elements to an Excel workbook.

        Parameters:
            elements (list[dict]): List of OSM elements from Overpass API.
            out_path (str): Output Excel file path.

        Returns:
            None
        """

        # Create a new Excel workbook and worksheet.
        wb = Workbook()
        ws = wb.active
        ws.title = "Overpass Results"

        # Write header row.
        ws.append(cls.HEADERS)

        # Write one row per OSM element.
        for element in elements:
            ws.append(cls.extract_row(element))

        # Freeze the top row for easier viewing in Excel.
        ws.freeze_panes = "A2"

        # Auto-adjust column widths based on content length.
        for col_idx, col_name in enumerate(cls.HEADERS, start=1):
            max_len = len(col_name)

            for row in ws.iter_rows(
                min_row=2,
                min_col=col_idx,
                max_col=col_idx,
                values_only=True
            ):
                val = row[0]
                if val is not None:
                    max_len = max(max_len, len(str(val)))

            # Limit maximum width to avoid extremely wide columns.
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        # Ensure output directory exists.
        os.makedirs(os.path.dirname(out_path), exist_ok=True)

        # Remove existing file before saving new workbook.
        if os.path.exists(out_path):
            os.remove(out_path)

        # Save workbook.
        wb.save(out_path)

        print(f"Saved: {out_path} rows: {ws.max_row - 1}")